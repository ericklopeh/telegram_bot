"""Generación de autorizaciones (Excel y PDF) con openpyxl y pypdf/reportlab."""

import json
import logging
import os
import re
import uuid
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
from sqlalchemy.orm import Session

from app.config import get_settings
from app.domain import constants as C
from app.models.authorization_job import AuthorizationJob
from app.models.case import Case
from app.models.document import Document
from app.models.talon_review import TalonReview
from app.services.document_service import DocumentService, StoredIncomingFile
from app.services.pdf_orden_service import generar_orden_snte_pdf, DatosOrden

log = logging.getLogger(__name__)


class TemplateNotFoundError(Exception):
    pass


def top_left_merge(ws, cell_addr: str) -> str:
    r, c = coordinate_to_tuple(cell_addr)
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
            return f"{get_column_letter(rng.min_col)}{rng.min_row}"
    return cell_addr


def escribir_seguro(ws, celda, valor, font=None, alignment=None, number_format=None):
    celda = top_left_merge(ws, celda)
    c = ws[celda]
    c.value = valor
    if font is not None:
        c.font = font
    if alignment is not None:
        c.alignment = alignment
    if number_format is not None:
        c.number_format = number_format


def arreglar_formula_final_venta(ws):
    ws["G13"].value = (
        "=INDEX(Hoja2!A2:ZZ40000,"
        "MATCH(Hoja1!F13,Hoja2!A2:A40000,0),"
        "MATCH(Hoja1!F14,Hoja2!A1:ZZ1,0))"
    )


def calcular_qna_final(qna_inicial: str, plazo: int) -> str:
    m = re.fullmatch(r"\s*(\d{1,2})-(\d{4})\s*", qna_inicial or "")
    if not m:
        return ""
    nn, yyyy = int(m.group(1)), int(m.group(2))
    total = (yyyy * 24 + (nn - 1)) + (plazo - 1)
    return f"{(total % 24) + 1:02d}-{total // 24}"


class AuthorizationService:
    def __init__(self, db: Session):
        self.db = db
        self.settings = get_settings()
        self.doc_service = DocumentService()

    def generate_for_case(self, case_id: int, form_data: dict, action_user: str) -> list[Document]:
        case = self.db.query(Case).filter(Case.id == case_id).first()
        if not case:
            raise ValueError(f"Caso {case_id} no encontrado")

        upload_dir = self.doc_service.pedido_evidencias_dir(case)
        os.makedirs(upload_dir, exist_ok=True)

        cliente = form_data.get("nombre", case.client_name).strip()
        folio = form_data.get("folio", case.official_folio or case.temp_folio or "").strip()
        plazo = int(form_data.get("plazo_qnas") or 72)
        qna_inicial = form_data.get("qna_inicial", "").strip()
        monto_total = form_data.get("monto_total")
        monto_total = float(monto_total) if monto_total else 0.0

        # Si el monto total es 0, lo calculamos de los productos
        if monto_total == 0.0:
            for i in range(1, 6):
                pv = form_data.get(f"prod_{i}_precio")
                desc = form_data.get(f"prod_{i}_descuento")
                if pv:
                    monto_total += float(pv) - (float(desc) if desc else 0.0)

        qna_final = calcular_qna_final(qna_inicial, plazo)
        descuento_qna = form_data.get("descuento_qna", "").strip()
        if not descuento_qna and plazo > 0:
            descuento_qna = f"{(monto_total / plazo):.2f}"

        # 1. Generar Excel Maestro
        doc_excel = self._generar_excel_maestro(case, form_data, cliente, folio, plazo, qna_inicial, monto_total, upload_dir)

        # 2. Generar PDF SNTE
        doc_pdf = self._generar_pdf_snte(case, form_data, cliente, folio, plazo, qna_inicial, qna_final, descuento_qna, upload_dir)

        self.db.commit()
        return [doc_excel, doc_pdf]

    def _generar_excel_maestro(self, case: Case, form_data: dict, cliente: str, folio: str, plazo: int, qna_inicial: str, monto_total: float, upload_dir: Path) -> Document:
        template_path = "storage/templates/plantilla_master_autorizaciones.xlsx"
        if not os.path.exists(template_path):
            raise TemplateNotFoundError(f"Falta plantilla Excel en: {template_path}")

        wb = openpyxl.load_workbook(template_path, data_only=False)
        ws = wb["Hoja1"] if "Hoja1" in wb.sheetnames else wb.active
        try:
            wb._external_links = []
        except Exception:
            pass

        estilo = Font(name="Aptos Narrow", size=12, bold=True, color="000000")
        alineacion = Alignment(horizontal="center", vertical="center")

        def escribir(celda, valor, centrar=True):
            escribir_seguro(ws, celda, valor, font=estilo, alignment=alineacion if centrar else None)

        escribir("D1", form_data.get("tel_part", ""))
        escribir("B2", cliente, centrar=False)
        escribir("E2", form_data.get("rfc", ""))
        try:
            fecha_obj = datetime.strptime(form_data.get("fecha_venta", ""), "%d/%m/%Y")
            escribir_seguro(ws, "G2", fecha_obj, font=estilo, alignment=alineacion, number_format="dd/mm/yyyy")
        except Exception:
            pass

        escribir("A11", form_data.get("observaciones", ""), centrar=False)
        escribir("B15", case.seller_name, centrar=False)
        escribir("D11", folio)
        escribir("F11", form_data.get("semana", ""))
        escribir_seguro(ws, "F13", qna_inicial, font=estilo, alignment=alineacion, number_format="@")
        escribir_seguro(ws, "F14", plazo, font=estilo, alignment=alineacion, number_format="0")

        # Escribir Productos en las Filas 4 a 8
        hay_productos = False
        for i in range(1, 6):
            r = 3 + i  # Fila 4, 5, 6, 7, 8
            prod = form_data.get(f"prod_{i}_nombre", "").strip()
            if not prod:
                continue
            hay_productos = True
            trans = form_data.get(f"prod_{i}_trans", "")
            cred = form_data.get(f"prod_{i}_credito", "")
            pv = form_data.get(f"prod_{i}_precio", "")
            desc = form_data.get(f"prod_{i}_descuento", "")
            tipo = form_data.get(f"prod_{i}_tipo", "")
            
            escribir(f"A{r}", prod, centrar=False)
            if trans: escribir_seguro(ws, f"C{r}", float(trans), font=estilo, alignment=alineacion, number_format='"$"#,##0.00')
            if cred: escribir_seguro(ws, f"D{r}", float(cred), font=estilo, alignment=alineacion, number_format='"$"#,##0.00')
            if pv: escribir_seguro(ws, f"E{r}", float(pv), font=estilo, alignment=alineacion, number_format='"$"#,##0.00')
            if desc: escribir_seguro(ws, f"F{r}", float(desc), font=estilo, alignment=alineacion, number_format='"$"#,##0.00')
            if tipo: escribir(f"G{r}", tipo)

        # Si no capturaron productos en la tabla, rellenamos 1 fila genérica con el monto total
        if not hay_productos:
            tipo_venta = "MUEBLE" if str(case.order_type).lower() == "mueble" else str(case.order_type).upper()
            escribir("A4", f"VENTA DE {tipo_venta}", centrar=False)
            escribir_seguro(ws, "D4", monto_total, font=estilo, alignment=alineacion, number_format='"$"#,##0.00') # Crédito
            escribir_seguro(ws, "E4", monto_total, font=estilo, alignment=alineacion, number_format='"$"#,##0.00') # Precio Venta
            escribir("G4", "NUEVA")

        # Escribir Descuento en F12
        descuento_val = form_data.get("descuento_qna")
        if not descuento_val and plazo > 0:
            descuento_val = monto_total / plazo
        if descuento_val:
            escribir_seguro(ws, "F12", float(descuento_val), font=estilo, alignment=alineacion, number_format='"$"#,##0.00')

        # Escribir Monto Total en D12
        escribir_seguro(ws, "D12", monto_total, font=estilo, alignment=alineacion, number_format='"$"#,##0.00')

        # Escribir Qna Final calculada en G14 (debajo de FINAL)
        qna_fin = calcular_qna_final(qna_inicial, plazo)
        escribir_seguro(ws, "G14", qna_fin, font=estilo, alignment=alineacion, number_format="@")

        filename = f"{uuid.uuid4()}_autorizacion.xlsx"
        abs_path = os.path.join(upload_dir, filename)
        wb.save(abs_path)

        job = AuthorizationJob(
            case_id=case.id,
            template_name=os.path.basename(template_path),
            output_path=abs_path,
            generation_status="success"
        )
        self.db.add(job)

        stored_file = StoredIncomingFile(
            stored_filename=filename,
            file_path=abs_path,
            original_filename=f"Autorizacion_{cliente}.xlsx",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        # Re-usamos el DOC_AUTORIZACION_SNTE para ambos, o podríamos registrar el excel como 'pedido' y el pdf como 'autorizacion'. 
        # Mantendremos DOC_AUTORIZACION_SNTE para simplificar.
        doc, _ = self.doc_service.register_pedido_document_upload(self.db, case, C.DOC_AUTORIZACION_SNTE, stored_file)
        return doc

    def _generar_pdf_snte(self, case: Case, form_data: dict, cliente: str, folio: str, plazo: int, qna_inicial: str, qna_final: str, descuento_qna: str, upload_dir: Path) -> Document:
        template_path = "storage/templates/plantilla_orden_snte.pdf"
        if not os.path.exists(template_path):
            raise TemplateNotFoundError(f"Falta plantilla PDF SNTE en: {template_path}")

        datos = DatosOrden(
            nombre=cliente.upper(),
            rfc=form_data.get("rfc", "").upper(),
            categoria=form_data.get("categoria", "").upper(),
            domicilio=form_data.get("domicilio", "").upper(),
            tel_part=form_data.get("tel_part", ""),
            tel_celular=form_data.get("tel_celular", ""),
            correo=form_data.get("correo", ""),
            fecha_venta=form_data.get("fecha_venta", ""),
            qna_inicial=qna_inicial,
            qna_final=qna_final,
            descuento_qna=descuento_qna,
            plazo_qnas=str(plazo),
            folio=folio,
        )

        nombre_safe = "".join(c for c in cliente if c.isalnum() or c == " ").replace(" ", "_")
        filename = f"{uuid.uuid4()}_{folio}_{nombre_safe}_SNTE.pdf"
        abs_path = os.path.join(upload_dir, filename)

        generar_orden_snte_pdf(template_path, abs_path, datos)

        job = AuthorizationJob(
            case_id=case.id,
            template_name=os.path.basename(template_path),
            output_path=abs_path,
            generation_status="success"
        )
        self.db.add(job)

        stored_file = StoredIncomingFile(
            stored_filename=filename,
            file_path=abs_path,
            original_filename=f"{folio}_{nombre_safe}_SNTE.pdf",
            mime_type="application/pdf"
        )
        doc, _ = self.doc_service.register_pedido_document_upload(self.db, case, C.DOC_ORDEN_SNTE_PDF, stored_file)
        return doc
