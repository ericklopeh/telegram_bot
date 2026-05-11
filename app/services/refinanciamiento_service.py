"""Generación de autorizaciones de refinanciamiento (Excel + PDF SNTE)."""

import logging
import os
import re
import uuid
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
from sqlalchemy.orm import Session

from app.config import get_settings
from app.domain import constants as C
from app.models.authorization_job import AuthorizationJob
from app.models.case import Case
from app.models.document import Document
from app.services.document_service import DocumentService, StoredIncomingFile
from app.services.pdf_orden_service import DatosOrden, generar_orden_snte_pdf

log = logging.getLogger(__name__)

TEMPLATE_REFI = "storage/templates/plantilla_refinanciamiento.xlsx"


class TemplateNotFoundError(Exception):
    """Plantilla requerida no encontrada en disco."""


# ---------------------------------------------------------------------------
# Helpers de celda (copiados de authorization_service para no acoplarnos)
# ---------------------------------------------------------------------------

def _top_left_merge(ws, cell_addr: str) -> str:
    """Devuelve la celda top-left si la dirección cae dentro de un merge."""
    r, c = coordinate_to_tuple(cell_addr)
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
            return f"{get_column_letter(rng.min_col)}{rng.min_row}"
    return cell_addr


def _escribir_seguro(ws, celda: str, valor, font=None, alignment=None, number_format=None) -> None:
    """Escribe un valor en la celda correcta aunque esté fusionada."""
    celda = _top_left_merge(ws, celda)
    c = ws[celda]
    c.value = valor
    if font is not None:
        c.font = font
    if alignment is not None:
        c.alignment = alignment
    if number_format is not None:
        c.number_format = number_format


def _calcular_qna_final(qna_inicial: str, plazo: int) -> str:
    """Calcula la quincena final a partir de la inicial y el plazo."""
    m = re.fullmatch(r"\s*(\d{1,2})-(\d{4})\s*", qna_inicial or "")
    if not m:
        return ""
    nn, yyyy = int(m.group(1)), int(m.group(2))
    total = (yyyy * 24 + (nn - 1)) + (plazo - 1)
    return f"{(total % 24) + 1:02d}-{total // 24}"


# ---------------------------------------------------------------------------
# Servicio principal
# ---------------------------------------------------------------------------

class RefinanciamientoService:
    """Orquesta la generación de documentos para refinanciamientos SNTE."""

    def __init__(self, db: Session) -> None:
        self.db = db
        self.settings = get_settings()
        self.doc_service = DocumentService()

    # ------------------------------------------------------------------
    # API pública
    # ------------------------------------------------------------------

    def generate_for_case(
        self,
        case_id: int,
        form_data: dict,
        action_user: str,
    ) -> list[Document]:
        """Genera Excel de refinanciamiento + PDF Orden SNTE para un caso.

        Retorna [doc_excel, doc_pdf].
        Deja ambos con upload_status = PENDING_UPLOAD (sin subir a SharePoint).
        """
        case = self.db.query(Case).filter(Case.id == case_id).first()
        if not case:
            raise ValueError(f"Caso {case_id} no encontrado")

        upload_dir = self.doc_service.pedido_evidencias_dir(case)
        os.makedirs(upload_dir, exist_ok=True)

        # Datos comunes derivados del form
        cliente = form_data.get("nombre", case.client_name or "").strip()
        folio = form_data.get("folio", case.official_folio or case.temp_folio or "").strip()
        plazo = int(form_data.get("plazo_qnas") or 72)
        qna_inicial = form_data.get("qna_inicial", "").strip()

        # Monto total: capturado o suma de precios de productos
        monto_total_raw = form_data.get("monto_total")
        monto_total = float(monto_total_raw) if monto_total_raw else 0.0
        if monto_total == 0.0:
            for i in range(1, 6):
                pv = form_data.get(f"prod_{i}_precio")
                if pv:
                    monto_total += float(pv)

        qna_final = _calcular_qna_final(qna_inicial, plazo)

        descuento_qna = form_data.get("descuento_qna", "").strip()
        if not descuento_qna and plazo > 0 and monto_total > 0:
            descuento_qna = f"{(monto_total / plazo):.2f}"

        # 1. Excel
        doc_excel = self._generar_excel_refi(
            case, form_data, cliente, folio, plazo,
            qna_inicial, monto_total, upload_dir,
        )

        # 2. PDF Orden SNTE
        doc_pdf = self._generar_pdf_snte(
            case, form_data, cliente, folio, plazo,
            qna_inicial, qna_final, descuento_qna, upload_dir,
        )

        self.db.commit()
        return [doc_excel, doc_pdf]

    # ------------------------------------------------------------------
    # Excel de refinanciamiento
    # ------------------------------------------------------------------

    def _generar_excel_refi(
        self,
        case: Case,
        form_data: dict,
        cliente: str,
        folio: str,
        plazo: int,
        qna_inicial: str,
        monto_total: float,
        upload_dir: Path,
    ) -> Document:
        if not os.path.exists(TEMPLATE_REFI):
            raise TemplateNotFoundError(f"Falta plantilla Excel en: {TEMPLATE_REFI}")

        wb = openpyxl.load_workbook(TEMPLATE_REFI, data_only=False)
        ws = wb["Hoja1"] if "Hoja1" in wb.sheetnames else wb.active

        # Limpiar external links cacheados
        try:
            wb._external_links = []
        except Exception:
            pass

        estilo = Font(name="Aptos Narrow", size=12, bold=True, color="000000")
        alineacion = Alignment(horizontal="center", vertical="center")
        fmt_money = '"$"#,##0.00'
        fmt_text = "@"
        fmt_int = "0"

        def escribir(celda: str, valor, centrar: bool = True) -> None:
            _escribir_seguro(
                ws, celda, valor,
                font=estilo,
                alignment=alineacion if centrar else None,
            )

        def escribir_money(celda: str, valor: float) -> None:
            _escribir_seguro(ws, celda, float(valor), font=estilo, alignment=alineacion, number_format=fmt_money)

        # ── Encabezado ──────────────────────────────────────────────────
        escribir("D1", form_data.get("tel_part", ""))
        escribir("B2", cliente, centrar=False)
        escribir("E2", form_data.get("rfc", ""))
        try:
            fecha_obj = datetime.strptime(form_data.get("fecha_venta", ""), "%d/%m/%Y")
            _escribir_seguro(ws, "G2", fecha_obj, font=estilo, alignment=alineacion, number_format="dd/mm/yyyy")
        except Exception:
            pass  # fecha opcional

        # ── Tabla de productos (filas 4-8 → hasta 5 productos) ──────────
        for i in range(1, 6):
            row = 3 + i  # filas 4, 5, 6, 7, 8
            nombre = form_data.get(f"prod_{i}_nombre", "").strip()
            codigo = form_data.get(f"prod_{i}_codigo", "").strip()
            trans_raw = form_data.get(f"prod_{i}_trans")
            precio_raw = form_data.get(f"prod_{i}_precio")

            if nombre:
                escribir(f"A{row}", nombre, centrar=False)
            if codigo:
                escribir(f"B{row}", codigo, centrar=False)
            if trans_raw:
                escribir_money(f"C{row}", float(trans_raw))
            if precio_raw:
                escribir_money(f"E{row}", float(precio_raw))

            # Tipo de venta: siempre REFINANCIADA cuando hay producto
            if nombre or precio_raw:
                escribir(f"G{row}", "REFINANCIADA")

        # ── Saldos a reestructurar (filas 14-18 → hasta 5 saldos) ──────
        for i in range(1, 6):
            row = 13 + i  # filas 14, 15, 16, 17, 18
            folio_refi = form_data.get(f"refi_folio_{i}", "").strip()
            desc_refi = form_data.get(f"refi_descuento_{i}")
            saldo_refi = form_data.get(f"refi_saldo_{i}")

            if folio_refi:
                _escribir_seguro(ws, f"C{row}", folio_refi, font=estilo, alignment=alineacion, number_format=fmt_text)
            if desc_refi:
                escribir_money(f"E{row}", float(desc_refi))
            if saldo_refi:
                escribir_money(f"F{row}", float(saldo_refi))

        # ── Monto venta nueva ────────────────────────────────────────────
        monto_vta_nueva = form_data.get("monto_vta_nueva")
        if monto_vta_nueva:
            escribir_money("F22", float(monto_vta_nueva))

        # ── Datos de descuento / crédito ─────────────────────────────────
        escribir("D23", folio)
        semana_raw = form_data.get("semana", "")
        if semana_raw:
            try:
                _escribir_seguro(ws, "F23", int(semana_raw), font=estilo, alignment=alineacion, number_format=fmt_int)
            except ValueError:
                escribir("F23", semana_raw)

        if monto_total:
            escribir_money("C25", monto_total)

        descuento_qna_raw = form_data.get("descuento_qna")
        if descuento_qna_raw:
            escribir_money("E25", float(descuento_qna_raw))

        _escribir_seguro(ws, "F25", qna_inicial, font=estilo, alignment=alineacion, number_format=fmt_text)
        _escribir_seguro(ws, "F26", plazo, font=estilo, alignment=alineacion, number_format=fmt_int)

        # ── Vendedor ─────────────────────────────────────────────────────
        vendedor = (case.seller_name or form_data.get("vendedor", "")).strip()
        if vendedor:
            escribir("B27", vendedor, centrar=False)

        # ── Forzar recálculo al abrir ────────────────────────────────────
        try:
            wb.calculation.calcMode = "auto"
            wb.calculation.fullCalcOnLoad = True
        except Exception:
            pass

        # ── Guardar ──────────────────────────────────────────────────────
        filename = f"{uuid.uuid4()}_refinanciamiento.xlsx"
        abs_path = str(upload_dir / filename)
        wb.save(abs_path)

        # ── Registrar AuthorizationJob ───────────────────────────────────
        job = AuthorizationJob(
            case_id=case.id,
            template_name=os.path.basename(TEMPLATE_REFI),
            output_path=abs_path,
            generation_status="success",
        )
        self.db.add(job)

        # ── Registrar Document (PENDING_UPLOAD) ──────────────────────────
        nombre_safe = "".join(c for c in cliente if c.isalnum() or c == " ").replace(" ", "_")
        stored_file = StoredIncomingFile(
            stored_filename=filename,
            file_path=abs_path,
            original_filename=f"Refinanciamiento_{nombre_safe}.xlsx",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        doc, _ = self.doc_service.register_pedido_document_upload(
            self.db, case, C.DOC_AUTORIZACION_REFI, stored_file
        )
        return doc

    # ------------------------------------------------------------------
    # PDF Orden SNTE (mismo formato que Venta Nueva)
    # ------------------------------------------------------------------

    def _generar_pdf_snte(
        self,
        case: Case,
        form_data: dict,
        cliente: str,
        folio: str,
        plazo: int,
        qna_inicial: str,
        qna_final: str,
        descuento_qna: str,
        upload_dir: Path,
    ) -> Document:
        template_path = "storage/templates/plantilla_orden_snte.pdf"
        if not os.path.exists(template_path):
            raise TemplateNotFoundError(f"Falta plantilla PDF SNTE en: {template_path}")

        datos = DatosOrden(
            nombre=cliente.upper(),
            rfc=form_data.get("rfc", "").upper(),
            categoria=form_data.get("categoria", "").upper(),
            domicilio=form_data.get("domicilio", "").upper(),
            tel_part=form_data.get("tel_part", ""),
            tel_celular=form_data.get("tel_celular", ""),
            correo=form_data.get("correo", ""),
            fecha_venta=form_data.get("fecha_venta", ""),
            qna_inicial=qna_inicial,
            qna_final=qna_final,
            descuento_qna=descuento_qna,
            plazo_qnas=str(plazo),
            folio=folio,
        )

        nombre_safe = "".join(c for c in cliente if c.isalnum() or c == " ").replace(" ", "_")
        filename = f"{uuid.uuid4()}_{folio}_{nombre_safe}_REFI_SNTE.pdf"
        abs_path = str(upload_dir / filename)

        generar_orden_snte_pdf(template_path, abs_path, datos)

        job = AuthorizationJob(
            case_id=case.id,
            template_name=os.path.basename(template_path),
            output_path=abs_path,
            generation_status="success",
        )
        self.db.add(job)

        stored_file = StoredIncomingFile(
            stored_filename=filename,
            file_path=abs_path,
            original_filename=f"{folio}_{nombre_safe}_REFI_SNTE.pdf",
            mime_type="application/pdf",
        )
        doc, _ = self.doc_service.register_pedido_document_upload(
            self.db, case, C.DOC_ORDEN_SNTE_PDF, stored_file
        )
        return doc
